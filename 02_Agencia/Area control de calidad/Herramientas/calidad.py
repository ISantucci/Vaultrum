#!/usr/bin/env python3
"""Vaultrum - Area de Control de Calidad - medicion del gate (v1).

Mide un QA instrumentado contra las seis leyes de la verificacion y calcula el
veredicto. El veredicto declarado tiene que coincidir con el medido.

  Ley 1  nada se verifica sin version congelada
  Ley 2  antes del pase, la build se acepta o se rechaza
  Ley 3  un hallazgo se reproduce, o se declara intermitente
  Ley 4  nada se cierra sin reverificar
  Ley 5  un pase declara lo que no ejecuto
  Ley 6  un riesgo aceptado tiene dueno con nombre

  python3 calidad.py <archivo|carpeta>                      informe
  python3 calidad.py <archivo|carpeta> --verificar          solo veredicto (exit 1 si falla)
  python3 calidad.py <archivo> --planilla <archivo.xlsx>    cruza con el registro de defectos

Las excepciones se declaran en Herramientas/excepciones.txt: ruta | ley | razon.
Lo que no esta declarado ahi, falla.
"""
import os, re, sys

SEVERIDADES = ('bloqueante', 'critico', 'mayor', 'menor', 'trivial')
ESTADOS     = ('abierto', 'diferido', 'cerrado')
VEREDICTOS  = ('GO', 'CONDITIONAL GO', 'NO-GO')
PERFILES    = ('ligero', 'estandar', 'completo')
LEYES = {
    'A':  'Contrato del QA - alcance, insumo y perfil',
    'L1': 'Ley 1 - version congelada',
    'L2': 'Ley 2 - verificacion de build',
    'L3': 'Ley 3 - hallazgo reproducible',
    'L4': 'Ley 4 - nada se cierra sin reverificar',
    'L5': 'Ley 5 - se declara lo no ejecutado',
    'L6': 'Ley 6 - riesgo aceptado con dueno',
    'V':  'Veredicto - lo declarado coincide con lo medido',
}


# ------------------------------------------------------------------ lectura
def bloques(txt):
    """Devuelve {etiqueta: [lineas]} de los bloques cercados ```qa-*."""
    out, actual, buf = {}, None, []
    for linea in txt.split('\n'):
        s = linea.strip()
        if s.startswith('```'):
            info = s[3:].strip().lower()
            if actual is not None:
                out.setdefault(actual, []).extend(buf)
                actual, buf = None, []
            elif info.startswith('qa-'):
                actual, buf = info, []
            continue
        if actual is not None and s:
            buf.append(s)
    return out


ANOTACION = re.compile(r'\s*\([^)]*\)\s*$')


def campos(lineas):
    """clave  valor  -> dict (primer token es la clave).

    Una anotacion entre parentesis al final del valor es documentacion del
    formato, no parte del valor: `perfil estandar (ligero | estandar | completo)`.
    """
    d = {}
    for l in lineas:
        p = l.split(None, 1)
        if len(p) == 2:
            d[p[0].strip().lower()] = ANOTACION.sub('', p[1].strip()).strip()
        elif p:
            d[p[0].strip().lower()] = ''
    return d


def filas(lineas):
    return [[c.strip() for c in l.split('|')] for l in lineas if '|' in l]


def cargar_excepciones(raiz):
    ruta = os.path.join(raiz, 'excepciones.txt')
    exc = {}
    if not os.path.exists(ruta):
        return exc
    for l in open(ruta, encoding='utf-8', errors='replace'):
        l = l.split('#')[0].strip()
        if not l or '|' not in l:
            continue
        p = [x.strip() for x in l.split('|')]
        if len(p) >= 3:
            exc.setdefault(p[0].replace('\\', '/'), {})[p[1]] = p[2]
    return exc


# ------------------------------------------------------------------ planilla
def leer_planilla(ruta):
    """Devuelve {id: {campo: valor}} de la hoja Bugs, o None si no se puede leer."""
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
    except Exception:
        return None
    if 'Bugs' not in wb.sheetnames:
        return None
    ws = wb['Bugs']
    cab, out = None, {}
    for fila in ws.iter_rows(values_only=True):
        if cab is None:
            if fila and fila[0] and str(fila[0]).strip() == 'Bug ID':
                cab = [str(c).strip() if c else '' for c in fila]
            continue
        if not fila or not fila[0]:
            continue
        d = {cab[i]: (str(v).strip() if v is not None else '')
             for i, v in enumerate(fila) if i < len(cab)}
        out[str(fila[0]).strip().upper()] = d
    return out


# ------------------------------------------------------------------ medicion
def planilla_vecina(ruta):
    """La planilla del proyecto vive al lado del QA: se busca sola."""
    d = os.path.dirname(os.path.abspath(ruta))
    for f in sorted(os.listdir(d)) if os.path.isdir(d) else []:
        if f.lower().endswith('.xlsx') and not f.startswith('~$'):
            return os.path.join(d, f)
    return None


def medir(ruta, planilla=None):
    txt = open(ruta, encoding='utf-8', errors='replace').read()
    b = bloques(txt)
    fallas, notas = [], []

    alcance = campos(b.get('qa-alcance', []))
    perfil = alcance.get('perfil', '').lower()
    if perfil not in PERFILES:
        fallas.append(('A', 'perfil no declarado o invalido (%s)' % (perfil or 'vacio')))
        perfil = 'estandar'
    if not alcance.get('insumo'):
        fallas.append(('A', 'no declara su insumo (EJ-XXX.n o TL-XXX)'))

    # --- Ley 1: version congelada
    build = campos(b.get('qa-build', []))
    if not b.get('qa-build'):
        fallas.append(('L1', 'falta el bloque qa-build'))
    else:
        for campo in ('build', 'commit', 'plataforma', 'entorno'):
            v = build.get(campo, '')
            if not v or v.startswith('<'):
                fallas.append(('L1', 'qa-build sin %s' % campo))
        if build.get('congelada', '').lower() not in ('si', 'sí', 'yes'):
            fallas.append(('L1', 'la version no esta declarada congelada'))

    # --- Ley 2: verificacion de build
    humo = campos(b.get('qa-humo', []))
    resultado_humo = humo.get('resultado', '').lower()
    if not b.get('qa-humo'):
        fallas.append(('L2', 'falta el bloque qa-humo'))
    elif resultado_humo not in ('aceptada', 'condicional', 'rechazada'):
        fallas.append(('L2', 'resultado de build invalido (%s)' % (resultado_humo or 'vacio')))

    # --- riesgo: el perfil se justifica ahi, asi que tiene que existir
    riesgos = filas(b.get('qa-riesgo', []))
    if not riesgos:
        fallas.append(('A', 'sin analisis de riesgo declarado: el perfil no se puede justificar'))
    for f in riesgos:
        if len(f) < 5:
            fallas.append(('A', 'riesgo mal formado (sistema | prob | impacto | deteccion | '
                                'exposicion | modo de falla): %s' % ' | '.join(f)))
            continue
        for i, etiqueta in ((1, 'probabilidad'), (2, 'impacto'), (3, 'deteccion')):
            if not re.fullmatch(r'[1-5]', f[i]):
                fallas.append(('A', '%s: %s fuera de escala 1-5 (%s)' % (f[0], etiqueta, f[i])))
        if len(f) > 5 and not re.fullmatch(r'[1-5]', f[4]):
            fallas.append(('A', '%s: exposicion fuera de escala 1-5 (%s)' % (f[0], f[4])))

    # --- defectos
    defectos = []
    for f in filas(b.get('qa-defectos', [])):
        if len(f) < 3:
            fallas.append(('L3', 'defecto mal formado: %s' % ' | '.join(f)))
            continue
        did, sev, est = f[0].upper(), f[1].lower(), f[2].lower()
        nota = f[3] if len(f) > 3 else ''
        if sev not in SEVERIDADES:
            fallas.append(('L3', '%s: severidad invalida (%s)' % (did, sev)))
        if est not in ESTADOS:
            fallas.append(('L3', '%s: estado invalido (%s)' % (did, est)))
        # --- Ley 4
        if est == 'cerrado' and not nota:
            fallas.append(('L4', '%s: cerrado sin reverificacion declarada' % did))
        defectos.append((did, sev, est, nota))

    # --- Ley 3 contra la planilla
    if planilla is not None:
        for did, sev, est, _ in defectos:
            d = planilla.get(did)
            if d is None:
                fallas.append(('L3', '%s: no existe en el registro de defectos' % did))
                continue
            for campo in ('Steps to Reproduce', 'Expected Result', 'Actual Result',
                          'Reproducibility', 'Build'):
                if not d.get(campo):
                    fallas.append(('L3', '%s: sin %s en el registro' % (did, campo)))
            if not d.get('Evidence / Link'):
                fallas.append(('L3', '%s: sin evidencia en el registro' % did))
    else:
        if perfil == 'completo':
            fallas.append(('L3', 'perfil completo sin planilla: reproducibilidad no medida'))
        else:
            notas.append('sin planilla: la reproducibilidad de los defectos no se midio, se declara')

    # --- Ley 4: regresion
    regresiones = filas(b.get('qa-regresion', []))
    regresion_falla = any(len(f) > 2 and f[2].lower() not in ('ok', 'na') for f in regresiones)
    if perfil != 'ligero':
        if not regresiones:
            fallas.append(('L4', 'perfil %s sin regresion declarada' % perfil))
        for f in regresiones:
            if len(f) < 3:
                fallas.append(('L4', 'regresion mal formada: %s' % ' | '.join(f)))

    # --- Ley 5: cobertura
    cob = filas(b.get('qa-cobertura', []))
    if len(cob) < 2:
        fallas.append(('L5', 'cobertura ausente o sin filas de datos'))
    else:
        cab = [c.lower() for c in cob[0]]
        libres = {i for i, c in enumerate(cab) if c in ('sistema', 'plataforma')}
        for fila in cob[1:]:
            if len(fila) != len(cab):
                fallas.append(('L5', '%s: la fila no tiene las columnas de la cabecera'
                               % (fila[0] if fila else '?')))
                continue
            for i, celda in enumerate(fila):
                if not celda:
                    fallas.append(('L5', '%s: celda vacia en %s' % (fila[0], cab[i])))
                elif i not in libres:
                    v = celda.lower()
                    if v.startswith('na'):
                        if ':' not in celda or not celda.split(':', 1)[1].strip():
                            fallas.append(('L5', '%s: "no aplica" sin razon en %s'
                                           % (fila[0], cab[i])))
                    elif v not in ('si', 'sí', 'no'):
                        fallas.append(('L5', '%s: valor invalido en %s (%s)'
                                       % (fila[0], cab[i], celda)))

    huecos = sum(1 for fila in cob[1:] for c in fila if c.lower() == 'no') if len(cob) > 1 else 0

    # --- Ley 6: riesgo aceptado
    aceptados = {}
    for f in filas(b.get('qa-aceptado', [])):
        if len(f) < 4 or not f[2] or not f[3]:
            fallas.append(('L6', 'aceptacion sin dueno o sin razon: %s' % ' | '.join(f)))
            continue
        aceptados[f[0].upper()] = (f[2], f[3])
    for did, sev, est, _ in defectos:
        if est in ('abierto', 'diferido') and did not in aceptados:
            fallas.append(('L6', '%s queda %s y nadie declaro hacerse cargo' % (did, est)))

    # --- veredicto medido
    bloqueante = any(sev == 'bloqueante' and est != 'cerrado' for _, sev, est, _ in defectos)
    abiertos = [d for d in defectos if d[2] != 'cerrado']
    if resultado_humo == 'rechazada' or bloqueante or regresion_falla:
        medido = 'NO-GO'
    elif abiertos or huecos or resultado_humo == 'condicional':
        medido = 'CONDITIONAL GO'
    else:
        medido = 'GO'

    declarado = ' '.join(b.get('qa-decision', [])).strip().upper()
    if declarado not in VEREDICTOS:
        fallas.append(('V', 'veredicto no declarado o invalido (%s)' % (declarado or 'vacio')))
    elif declarado != medido:
        fallas.append(('V', 'veredicto declarado %s, medido %s' % (declarado, medido)))

    return dict(ruta=ruta, perfil=perfil, fallas=fallas, notas=notas, defectos=defectos,
                abiertos=len(abiertos), huecos=huecos, humo=resultado_humo,
                medido=medido, declarado=declarado)


# ------------------------------------------------------------------ informe
def archivos(destino):
    if os.path.isfile(destino):
        return [destino]
    out = []
    for dp, dn, fn in os.walk(destino):
        dn[:] = [d for d in dn if not d.startswith('.')]
        for f in sorted(fn):
            if f.startswith('QA-') and f.endswith('.md'):
                out.append(os.path.join(dp, f))
    return out


def main():
    args = [a for a in sys.argv[1:]]
    verificar = '--verificar' in args
    planilla_ruta = None
    if '--planilla' in args:
        i = args.index('--planilla')
        if i + 1 < len(args):
            planilla_ruta = args[i + 1]
            del args[i:i + 2]
    destino = next((a for a in args if not a.startswith('--')), '.')

    planilla = leer_planilla(planilla_ruta) if planilla_ruta else None
    if planilla_ruta and planilla is None:
        print('AVISO: no se pudo leer la planilla (%s). Se mide solo el QA.' % planilla_ruta)

    rutas = archivos(destino)
    if not rutas:
        print('No se encontro ningun QA-*.md en %s' % destino)
        return 1

    exc = cargar_excepciones(os.path.dirname(os.path.abspath(__file__)))
    total_fallas = 0

    for r in rutas:
        p = planilla
        if p is None and not planilla_ruta:
            vecina = planilla_vecina(r)
            p = leer_planilla(vecina) if vecina else None
        m = medir(r, p)
        vivas = []
        for ley, detalle in m['fallas']:
            rel = r.replace('\\', '/')
            razon = next((v[ley] for k, v in exc.items()
                          if ley in v and (rel == k or rel.endswith('/' + k))), None)
            if razon:
                m['notas'].append('excepcion declarada %s: %s' % (ley, razon))
            else:
                vivas.append((ley, detalle))
        total_fallas += len(vivas)

        if not verificar:
            print('\n' + '=' * 70)
            print(os.path.basename(r))
            print('=' * 70)
            print('  perfil %s | humo %s | defectos %d (abiertos %d) | huecos de cobertura %d'
                  % (m['perfil'], m['humo'] or '?', len(m['defectos']), m['abiertos'], m['huecos']))
            print('\n  Leyes:')
            for k in sorted(LEYES):
                fall = [d for l, d in vivas if l == k]
                print('   %-45s %s' % (LEYES[k], 'EN LEY' if not fall else 'FUERA DE LEY (%d)' % len(fall)))
                for d in fall:
                    print('        - %s' % d)
            for n in m['notas']:
                print('   nota: %s' % n)
            print('\n  veredicto medido     %s' % m['medido'])
            print('  veredicto declarado  %s' % (m['declarado'] or '-'))
        else:
            for n in m['notas']:
                print('%s: nota: %s' % (os.path.basename(r), n))
            if vivas:
                print('%s: %s' % (os.path.basename(r),
                                  ' | '.join('%s (%s)' % (LEYES[l], d) for l, d in vivas)))

    if total_fallas:
        print('\nGATE FUERA DE LEY: %d hallazgo(s). El QA no cierra.' % total_fallas)
        return 1
    print('\nGATE EN LEY: version congelada, build verificada, nada cerrado sin '
          'reverificar, cobertura declarada y riesgo con dueno.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
